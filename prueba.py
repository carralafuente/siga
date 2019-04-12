from django.shortcuts import render
import openpyxl

def CargaCrear(request):

	if request.method == 'GET':
		return render(request, 'carga_app/carga.html')
	else:
		excel_file = request.FILES["excel_file"]
		wb = openpyxl.load_workbook(excel_file)
		worksheet = wb["facturacion_pruebas"]#Elige la hoja a descargar del excel

		excel_data = list()
		for row in worksheet.iter_rows():
			row_data = list()
			for cell in row:
				row_data.append(str(cell.value))
			excel_data.append(row_data)
		return render(request, 'carga_app/carga.html', {"excel_data":excel_data})


#-----------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------
import xlrd
	
def UploadFile(request):
	
	if request.method == 'GET':
		return render(request, 'upload_app/upload_carga.html')
	
	else:
		#Abrir el workbook y la hoja de trabajo
		excel_file = request.FILES["prueba_carga"]
		book = xlrd.open_workbook(excel_file)
		sheet = book.sheet_by_index(0)

		#Establecer la conexión con la base de datos
		database = sqlite3.connect("siga/db.sqlite3")

		#Abrir el cursor y pasarlo línea por línea
		cursor = database.cursor()

		#Crear el querySet para insertar los datos a la BD
		QuerySet = '''INSERT INTO Upload (fecha, serie, folio, cliente, razonsocial, rfc, estado, neto, descuentos, impuestos, total) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''

		#For loop para iterar sobre casa línea del archivo xls, empezando en la línea 2 para saltar los headers
		for r in range(1, sheet.nrows)
			fecha = sheet.cell(r,1).value
			serie = sheet.cell(r,2).value
			folio = sheet.cell(r,3).value
			cliente = sheet.cell(r,4).value
			razonsocial = sheet.cell(r,5).value
			rfc = sheet.cell(r,6).value
			estado = sheet.cell(r,7).value
			neto = sheet.cell(r,8).value
			descuentos = sheet.cell(r,9).value
			impuestos = sheet.cell(r,10).value
			total = sheet.cell(r,11).value

			#Asignar los valores de cada línea
			values = (fecha, serie, folio, cliente, razonsocial, rfc, estado, neto, descuentos, impuestos, total)

			#Ejecutar el QuerySet
			cursor.execute(QuerySet, values)
		#Cerrar el cursor
		cursor.close()

		#Subir la transacción
		database.commit()

		#Cerrar la base de datos
		''' database.close()'''
#----------------------------------------------------------------------------------------------------------------------------

		